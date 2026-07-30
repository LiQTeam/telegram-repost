"""
گزارشِ فیدبکِ ادمین به فیلترِ تبلیغات.

این ماژول دو چیز می‌سازه:
  ۱) متنِ نمایشِ داخلِ خودِ ربات (خلاصه + جزئیاتِ هر کانال) - از رویِ
     database.get_ad_feedback_channel_stats / get_ad_feedback_posts.
  ۲) فایلِ اکسلِ کامل (دو برگه: خلاصه + جزئیات) برایِ دانلود.

هیچ کوئریِ مستقیمی به دیتابیس نمی‌زنه؛ فقط از توابعِ عمومیِ database.py که
از قبل owner_user_id رو درست فیلتر می‌کنن استفاده می‌کنه - پس ایزوله‌بودنِ
آمارِ هر کاربر/ادمین این‌جا تضمین‌شده‌ست، نه این‌جا دوباره پیاده‌سازی شده.
"""
from __future__ import annotations

import json
from datetime import datetime
from html import escape as _esc
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .database import db
from .formatter import strip_html_tags
from .jdatetime_utils import format_jalali_datetime, now_jalali

_REASON_PREFIX = "🚩 مشکوک به تبلیغاتی: "


def _clean_reason(flag_reason: str | None) -> str:
    flag_reason = flag_reason or ""
    if flag_reason.startswith(_REASON_PREFIX):
        flag_reason = flag_reason[len(_REASON_PREFIX):]
    return flag_reason.strip() or "—"


def _snippet(row, limit: int = 160) -> str:
    body = (row["body_html"] or row["caption_html"] or "") if row else ""
    try:
        plain = strip_html_tags(body).strip()
    except Exception:
        plain = body
    plain = " ".join(plain.split())
    if len(plain) > limit:
        plain = plain[:limit].rstrip() + "…"
    return plain or "(بدونِ متن)"


def _parse_detail(row) -> dict:
    """جزئیاتِ ساختاریافته‌ی موتورِ فیلتر (JSON ذخیره‌شده در ستونِ ad_filter_detail)
    رو پارس می‌کنه. برایِ پست‌های قدیمی که این ستون رو ندارن (قبل از اضافه‌شدنِ
    این قابلیت) یا هر خطایِ پارسِ دیگه، دیکشنریِ خالی برمی‌گردونه - گزارش هیچ‌وقت
    نباید به‌خاطرِ یک ردیفِ ناقص کلاً خطا بده."""
    try:
        raw = row["ad_filter_detail"]
    except (IndexError, KeyError):
        raw = None
    if not raw:
        return {}
    try:
        d = json.loads(raw)
        return d if isinstance(d, dict) else {}
    except Exception:
        return {}


def _bool_fa(v) -> str:
    return "بله" if v else "خیر"


def _rule_verdict_fa(detail: dict) -> str:
    if "rule_engine_verdict" not in detail:
        # وقتی داورِ AI اصلاً صدا زده نشده، خودِ is_ad نهایی همون تصمیمِ موتورِ
        # قاعده‌محوره؛ ولی چون این‌جا فقط جزئیاتِ خامِ analyze() رو داریم نه
        # verdictِ نهایی، از رویِ score/threshold مستقیماً محاسبه می‌کنیم.
        score = detail.get("score")
        threshold = detail.get("threshold")
        if score is None or threshold is None:
            return "—"
        return "تبلیغ" if score >= threshold else "عادی"
    return "تبلیغ" if detail.get("rule_engine_verdict") else "عادی"


def _llm_verdict_fa(detail: dict, key: str, confidence_key: str | None = None) -> str:
    val = detail.get(key)
    if val not in ("AD", "SAFE"):
        return "—"
    label = "تبلیغ" if val == "AD" else "عادی"
    if confidence_key and detail.get(confidence_key):
        conf = detail[confidence_key]
        conf_fa = "بالا" if conf == "high" else ("پایین" if conf == "low" else conf)
        label += f" (اطمینان {conf_fa})"
    return label


def _llm_check_text(detail: dict, key: str) -> str:
    """متنِ خامِ توضیحِ داورِ AI (خطِ CHECK). اگه اون داور اصلاً صدا زده نشده
    (مثلاً چون داورِ اول با موتورِ قاعده هم‌نظرِ مطمئن بود و نیازی به داورِ دوم
    نبود)، «—» برمی‌گردونه، نه خالی/None که در اکسل عجیب به‌نظر برسه."""
    txt = (detail.get(key) or "").strip()
    return txt or "—"


def _rule_reason_fa(row, detail: dict) -> str:
    """دلیلِ متنیِ خودِ موتورِ قاعده‌محور (بدونِ نظرِ AI، بدونِ HTML). برایِ پست‌های
    تازه از detail['rule_reason_text'] میاد (متنِ خام)؛ برایِ پست‌های قدیمی‌تر که
    این کلید رو ندارن، از rowِ ['flag_reason'] (که بنرِ HTMLِ کامل - شاملِ نظرِ AI
    هم هست) با حذفِ تگ‌های HTML استفاده می‌شه تا چیزی از دست نره."""
    rule_text = (detail.get("rule_reason_text") or "").strip()
    if rule_text:
        return rule_text
    raw = row["flag_reason"] if "flag_reason" in row.keys() else ""
    try:
        plain = strip_html_tags(raw or "")
    except Exception:
        plain = raw or ""
    plain = _clean_reason(plain)
    return plain


def _jalali(created_at: str | None) -> str:
    """created_at از pending_posts رشته‌ی خامِ SQLite است (مثلِ
    '2026-07-20 14:03:11')؛ اگه پارس نشد، همون رشته‌ی خام برگردونده می‌شه تا
    گزارش هیچ‌وقت به‌خاطرِ یک تاریخِ عجیب کلاً خطا نده."""
    if not created_at:
        return "—"
    try:
        dt = datetime.strptime(created_at.split(".")[0], "%Y-%m-%d %H:%M:%S")
        return format_jalali_datetime(dt)
    except Exception:
        return created_at


# ===========================================================================
#  متنِ نمایشِ داخلِ ربات
# ===========================================================================
def overview_text(stats: list[dict]) -> str:
    if not stats:
        return (
            "📊 <b>آمارِ فیدبکِ فیلترِ تبلیغات</b>\n\n"
            "هنوز هیچ فیدبکی برایِ کانال‌های شما ثبت نشده.\n"
            "وقتی پستی به‌خاطرِ مشکوک بودن به تبلیغ به صفِ تایید بیفته، زیرِ "
            "پیش‌نمایشش دو دکمه‌ی «✅ درست بود» و «❌ اشتباه بود» دیده می‌شه؛ با "
            "زدنِ اون دکمه‌ها این آمار کم‌کم پر می‌شه."
        )
    total_correct = sum(s["correct"] for s in stats)
    total_incorrect = sum(s["incorrect"] for s in stats)
    total = total_correct + total_incorrect
    pct = round(100 * total_correct / total) if total else 0
    lines = [
        "📊 <b>آمارِ فیدبکِ فیلترِ تبلیغات</b>",
        f"مجموع: {total} فیدبک · ✅ {total_correct} درست · ❌ {total_incorrect} اشتباه · دقتِ کلی ≈ {pct}٪",
        "",
        "یک کانالِ مبدأ رو از لیستِ زیر انتخاب کن تا مقصد(ها)، دلیلِ تشخیصِ "
        "فیلتر و چند نمونه‌ی اخیرش رو ببینی؛ یا کلِ این گزارش رو به‌صورتِ "
        "فایلِ اکسل دریافت کن.",
    ]
    return "\n".join(lines)


def channel_detail_text(stat: dict, posts: list, sample: int = 8) -> str:
    dest_txt = "، ".join(stat["destinations"]) if stat["destinations"] else "— (بدونِ مقصدِ متصل)"
    lines = [
        f"📡 <b>{_esc(stat['channel_name'])}</b>",
        f"🎯 مقصد(ها): {_esc(dest_txt)}",
        "",
        f"✅ درست بود: <b>{stat['correct']}</b>",
        f"❌ اشتباه بود: <b>{stat['incorrect']}</b>",
        f"📈 دقتِ فیلتر برایِ این کانال: <b>{stat['accuracy']}٪</b> (از {stat['total']} فیدبک)",
    ]
    if posts:
        lines.append("")
        lines.append(f"🕓 <b>آخرین {min(sample, len(posts))} نمونه:</b>")
        for row in posts[:sample]:
            mark = "✅" if row["ad_feedback"] == "correct" else "❌"
            reason = _clean_reason(row["flag_reason"])
            snip = _snippet(row, 110)
            lines.append(f"\n{mark} «{_esc(snip)}»\n   ↳ دلیلِ تشخیصِ فیلتر: {_esc(reason)}")
    return "\n".join(lines)


# ===========================================================================
#  خروجیِ اکسل
# ===========================================================================
# فونت‌هایِ IRANSans (از فایل‌هایِ آپلودشده): برایِ رندرِ دقیقاً همین فونت‌ها،
# باید خودِ فایل‌هایِ .ttf (توی fonts/iransans/ همین پروژه) رویِ سیستمی که اکسل
# رو باز می‌کنه نصب شده باشن - اکسل فونت رو داخلِ خودِ xlsx امبد نمی‌کنه، فقط
# اسمش رو ذخیره می‌کنه؛ اگه نصب نباشه، ویندوز/آفیس خودش با فونتِ پیش‌فرض جایگزینش
# می‌کنه (بدونِ خطا، فقط شکلِ ظاهریِ متن فرق می‌کنه).
_FONT_BODY = "IRANSans Medium"      # متنِ ردیف‌های داده - وزنِ متوسط، خوانا برایِ جدولِ پرمتن
_FONT_BOLD = "IRANSans"             # bold=True روی این خانواده مستقیماً به IRANSans Bold می‌ره
_FONT_HEADER = "IRANSans Black"     # هدرها/تیتر - سنگین‌ترین وزن برایِ کنتراستِ بالا
_FONT_LIGHT = "IRANSans Light"      # زیرتیتر/توضیحِ کم‌رنگ

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(name=_FONT_HEADER, size=11, bold=True, color="FFFFFF")
_GROUP_FONT = Font(name=_FONT_HEADER, size=10, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name=_FONT_HEADER, size=15, bold=True, color="1F4E78")
_SUBTITLE_FONT = Font(name=_FONT_LIGHT, size=10, italic=True, color="666666")
_BASE_FONT = Font(name=_FONT_BODY, size=10)
_BOLD_FONT = Font(name=_FONT_BOLD, size=10, bold=True)
_CORRECT_FILL = PatternFill("solid", fgColor="E2EFDA")
_INCORRECT_FILL = PatternFill("solid", fgColor="FCE4E4")
_TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_ZEBRA_FILL = PatternFill("solid", fgColor="F5F8FC")
_THIN = Side(style="thin", color="B7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, ncols: int, *, group: bool = False) -> None:
    font = _GROUP_FONT if group else _HEADER_FONT
    fill = _GROUP_FILL if group else _HEADER_FILL
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = font
        c.fill = fill
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[row].height = 20 if group else 24


def _autofit(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_ad_feedback_workbook(
    owner_user_id: int | None = None,
    channel_id: int | None = None,
    scope_label: str = "",
) -> bytes | None:
    """فایلِ اکسلِ کامل رو می‌سازه و بایت‌های خامش رو برمی‌گردونه (برایِ ارسالِ
    مستقیم با send_document، بدون نوشتنِ فایلِ موقت رویِ دیسک).
    owner_user_id: نگاه کن به database._ad_feedback_owner_clause - همیشه
    دقیقاً محدود به همون یک مالک (هیچ‌وقت «همه با هم»).
    channel_id: اگه پر باشه، فقط همون یک کانال (خروجیِ تک‌کانالی).
    اگه هیچ فیدبکی توی محدوده‌ی درخواستی نباشه، None برمی‌گردونه."""
    stats = db.get_ad_feedback_channel_stats(owner_user_id=owner_user_id)
    if channel_id is not None:
        stats = [s for s in stats if s["channel_id"] == channel_id]
    if not stats:
        return None

    posts_by_channel = {
        s["channel_id"]: db.get_ad_feedback_posts(owner_user_id=owner_user_id, channel_id=s["channel_id"])
        for s in stats
    }

    wb = Workbook()

    # ---------------- برگه‌ی ۱: خلاصه ----------------
    ws = wb.active
    ws.title = "خلاصه"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A5"

    ncols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="📊 گزارشِ فیدبکِ فیلترِ تبلیغات")
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 28

    subtitle = f"تاریخِ تولیدِ گزارش: {format_jalali_datetime(now_jalali())}"
    if scope_label:
        subtitle = f"{scope_label}  ·  {subtitle}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle)
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = _CENTER

    headers = ["📡 کانالِ مبدأ", "🎯 کانال‌های مقصد", "✅ درست بود", "❌ اشتباه بود", "مجموع", "📈 دقت"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, ncols)

    r = header_row + 1
    total_correct = total_incorrect = 0
    for s in stats:
        ws.cell(row=r, column=1, value=s["channel_name"])
        ws.cell(row=r, column=2, value="، ".join(s["destinations"]) or "—")
        c_correct = ws.cell(row=r, column=3, value=s["correct"])
        c_incorrect = ws.cell(row=r, column=4, value=s["incorrect"])
        ws.cell(row=r, column=5, value=s["total"])
        ws.cell(row=r, column=6, value=f"{s['accuracy']}٪")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = _BASE_FONT
            cell.border = _BORDER
            cell.alignment = _WRAP_RIGHT if col in (1, 2) else _CENTER
        c_correct.fill = _CORRECT_FILL
        c_incorrect.fill = _INCORRECT_FILL
        total_correct += s["correct"]
        total_incorrect += s["incorrect"]
        r += 1

    total_all = total_correct + total_incorrect
    overall_pct = round(100 * total_correct / total_all) if total_all else 0
    ws.cell(row=r, column=1, value="🔢 جمعِ کل")
    ws.cell(row=r, column=2, value="")
    ws.cell(row=r, column=3, value=total_correct)
    ws.cell(row=r, column=4, value=total_incorrect)
    ws.cell(row=r, column=5, value=total_all)
    ws.cell(row=r, column=6, value=f"{overall_pct}٪")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = _BOLD_FONT
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL
        cell.alignment = _WRAP_RIGHT if col == 1 else _CENTER

    _autofit(ws, {1: 30, 2: 38, 3: 12, 4: 12, 5: 10, 6: 10})
    ws.sheet_properties.tabColor = "1F4E78"
    ws.auto_filter.ref = f"A{header_row}:F{r - 1}"

    # نمودارِ میله‌ای: مقایسه‌ی «درست بود» / «اشتباه بود» به‌ازایِ هر کانال، کنارِ
    # همون جدول - یک نگاهِ گرافیکیِ سریع قبل از رفتن سراغِ جزئیاتِ خام.
    if len(stats) >= 1:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "مقایسه‌ی فیدبک به‌ازایِ هر کانال"
        chart.y_axis.title = "تعدادِ فیدبک"
        chart.style = 10
        chart.height = 8
        chart.width = max(16, 2.2 * len(stats) + 6)
        data_ref = Reference(ws, min_col=3, max_col=4, min_row=header_row, max_row=r - 1)
        cats_ref = Reference(ws, min_col=1, max_col=1, min_row=header_row + 1, max_row=r - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "70AD47"   # ✅ سبز
        chart.series[1].graphicalProperties.solidFill = "C00000"   # ❌ قرمز
        ws.add_chart(chart, f"A{r + 2}")

    # ---------------- برگه‌ی ۲: جزئیات ----------------
    ws2 = wb.create_sheet("جزئیات")
    ws2.sheet_view.rightToLeft = True
    ws2.freeze_panes = "A3"

    # ردیفِ اول: سرگروه‌های بصری (برایِ جداکردنِ «اطلاعاتِ پایه» از «تحلیلِ موتورِ
    # قاعده‌محور» و «نظرِ داورانِ AI» - چون تعدادِ ستون‌ها زیاده، این گروه‌بندی
    # کمک می‌کنه چشم سریع‌تر بفهمه هر بخش مالِ کجاست).
    base_cols = (1, 6)     # کانال/مقصد/فیدبک/دلیل/متن/تاریخ
    rule_cols = (7, 12)    # نتیجه/امتیاز/کلیدواژه/منشن‌ولینک/مرزی/معافیت
    ai_cols = (13, 16)     # نتیجه+شرحِ داورِ اول، نتیجه+شرحِ داورِ دوم
    ncols2 = ai_cols[1]

    def _merge_group(cols: tuple[int, int], label: str) -> None:
        c0, c1 = cols
        if c1 > c0:
            ws2.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        ws2.cell(row=1, column=c0, value=label)

    _merge_group(base_cols, "📋 اطلاعاتِ پایه‌یِ پست")
    _merge_group(rule_cols, "🔍 تحلیلِ موتورِ قاعده‌محور")
    _merge_group(ai_cols, "🧠 نظرِ داورانِ هوشِ مصنوعی")
    _style_header_row(ws2, 1, ncols2, group=True)

    headers2 = [
        "📡 کانالِ مبدأ", "🎯 کانال‌های مقصد", "نتیجه‌ی فیدبکِ ادمین", "دلیلِ موتورِ قاعده‌محور", "متنِ پست", "🗓 تاریخ",
        "نتیجه‌ی موتور", "امتیاز/آستانه", "کلیدواژه‌های تطبیق‌یافته", "منشن/لینک",
        "مرزی بود؟", "معافیتِ کانفیگ/پروکسی؟",
        "نتیجه‌ی داورِ اول", "شرحِ داورِ اول", "نتیجه‌ی داورِ دوم", "شرحِ داورِ دوم",
    ]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=2, column=i, value=h)
    _style_header_row(ws2, 2, ncols2)

    channel_dest_map = {s["channel_id"]: ("، ".join(s["destinations"]) or "—") for s in stats}
    channel_name_map = {s["channel_id"]: s["channel_name"] for s in stats}

    _CENTER_COLS = {3, 6, 7, 8, 10, 11, 12, 13, 15}
    r2 = 3
    zebra_idx = 0
    for s in stats:
        for row in posts_by_channel.get(s["channel_id"], []):
            verdict = row["ad_feedback"]
            mark = "✅ درست بود" if verdict == "correct" else "❌ اشتباه بود"
            detail = _parse_detail(row)
            score = detail.get("score")
            threshold = detail.get("threshold")
            score_txt = f"{score}/{threshold}" if score is not None and threshold is not None else "—"
            keywords_txt = "، ".join(detail.get("keywords") or []) or "—"
            ment_link_txt = f"{detail.get('mentions', '—')}/{detail.get('links', '—')}"

            values = {
                1: channel_name_map.get(s["channel_id"], "—"),
                2: channel_dest_map.get(s["channel_id"], "—"),
                3: mark,
                4: _rule_reason_fa(row, detail),
                5: _snippet(row, 400),
                6: _jalali(row["created_at"]),
                7: _rule_verdict_fa(detail),
                8: score_txt,
                9: keywords_txt,
                10: ment_link_txt,
                11: _bool_fa(detail.get("borderline")) if detail else "—",
                12: _bool_fa(detail.get("config")) if detail else "—",
                13: _llm_verdict_fa(detail, "llm", "llm_confidence"),
                14: _llm_check_text(detail, "llm_check"),
                15: _llm_verdict_fa(detail, "llm2"),
                16: _llm_check_text(detail, "llm2_check"),
            }
            zebra = zebra_idx % 2 == 1
            for col in range(1, ncols2 + 1):
                cell = ws2.cell(row=r2, column=col, value=values[col])
                cell.font = _BASE_FONT
                cell.border = _BORDER
                cell.alignment = _CENTER if col in _CENTER_COLS else _WRAP_RIGHT
                if zebra:
                    cell.fill = _ZEBRA_FILL
            ws2.cell(row=r2, column=3).fill = _CORRECT_FILL if verdict == "correct" else _INCORRECT_FILL
            ws2.row_dimensions[r2].height = 46
            zebra_idx += 1
            r2 += 1

    ws2.sheet_properties.tabColor = "C00000"
    if r2 > 3:
        ws2.auto_filter.ref = f"A2:{get_column_letter(ncols2)}{r2 - 1}"

    _autofit(ws2, {
        1: 22, 2: 26, 3: 16, 4: 34, 5: 46, 6: 17,
        7: 13, 8: 12, 9: 26, 10: 11, 11: 10, 12: 15,
        13: 20, 14: 34, 15: 20, 16: 34,
    })

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
